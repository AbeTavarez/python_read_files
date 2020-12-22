import openpyxl
import os

# User if file is in a different directory
# os.chdir('/Users/<username>/Downloads')

# Load workbook by name
workbook = openpyxl.load_workbook('example.xlsx')

# Shows type
print(type(workbook))

# Grabs sheet by name
sheet = workbook['Sheet1']
print(sheet)

# returns array of sheets
list_sheet = workbook.sheetnames
print(list_sheet)

# Gets cell from a sheet
cell = sheet['A1']

# Print cell value
print(str(cell.value))
# or
print(str(sheet['A1'].value))
print(sheet['B1'].value)
# * It return the type that was formated on the excel file
print(sheet['C1'].value)

# returns cell object
print(sheet.cell(row=1, column=2))
# same as
print(sheet['B1'])

#
for i in range(1, 8):
    print(sheet.cell(row=i, column=2).value)
